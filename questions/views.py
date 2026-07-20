import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
import openpyxl
from accounts.decorators import teacher_required
from .models import Question, Choice, Subject, Topic
from .forms import QuestionForm, ChoiceFormSet, TrueFalseFormSet, SubjectForm, ImportQuestionForm
from .utils import import_questions_from_excel, import_questions_from_docx


@teacher_required
def question_bank_view(request):
    qs = Question.objects.select_related('subject', 'topic').prefetch_related('choices')
    subject_id = request.GET.get('subject', '')
    topic_id = request.GET.get('topic', '')
    q_type = request.GET.get('type', '')
    difficulty = request.GET.get('difficulty', '')
    q = request.GET.get('q', '')

    if subject_id:
        qs = qs.filter(subject_id=subject_id)
        # Chủ đề chỉ có hiệu lực khi đã chọn Môn học (topic thuộc môn)
        if topic_id:
            qs = qs.filter(topic_id=topic_id)
    else:
        topic_id = ''
    if q_type:
        qs = qs.filter(question_type=q_type)
    if difficulty:
        qs = qs.filter(difficulty=difficulty)
    if q:
        qs = qs.filter(content__icontains=q)

    # Danh sách chủ đề của môn đang chọn (để render dropdown)
    topics = Topic.objects.filter(subject_id=subject_id).order_by('name') if subject_id else []

    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'questions/bank.html', {
        'page_obj': page,
        'subjects': Subject.objects.all(),
        'topics': topics,
        'type_choices': Question.TYPE_CHOICES,
        'difficulty_choices': Question.DIFFICULTY_CHOICES,
        'filters': {'subject': subject_id, 'topic': topic_id, 'type': q_type, 'difficulty': difficulty, 'q': q},
    })


def _validate_choices(formset, q_type):
    """Kiểm tra số đáp án và số đáp án đúng sau khi formset hợp lệ. Trả về thông báo lỗi hoặc None."""
    n_total, n_correct = 0, 0
    for f in formset:
        cd = getattr(f, 'cleaned_data', None) or {}
        if cd.get('DELETE') or not (cd.get('content') or '').strip():
            continue
        n_total += 1
        if cd.get('is_correct'):
            n_correct += 1
    if n_total < 2:
        return 'Cần nhập ít nhất 2 đáp án.'
    if q_type == 'single' and n_correct != 1:
        return f'Trắc nghiệm 1 đáp án phải chọn ĐÚNG 1 đáp án đúng (bạn đang chọn {n_correct}).'
    if q_type == 'multiple' and n_correct < 2:
        return f'Trắc nghiệm nhiều đáp án phải chọn ÍT NHẤT 2 đáp án đúng (bạn đang chọn {n_correct}).'
    return None


def _save_true_false_choices(question, tf_correct):
    """Tạo/cập nhật 2 đáp án Đúng-Sai theo lựa chọn đáp án đúng ('true'/'false')."""
    existing = list(question.choices.order_by('order'))
    if len(existing) == 2:
        for c in existing:
            is_true_choice = c.content.strip().lower() in ('đúng', 'dung', 'true')
            c.is_correct = (is_true_choice == (tf_correct == 'true'))
            c.save(update_fields=['is_correct'])
    else:
        question.choices.all().delete()
        Choice.objects.create(question=question, label='Đ', content='Đúng', is_correct=(tf_correct == 'true'), order=1)
        Choice.objects.create(question=question, label='S', content='Sai', is_correct=(tf_correct == 'false'), order=2)


@teacher_required
def question_create_view(request):
    form = QuestionForm(request.POST or None, request.FILES or None)
    formset = ChoiceFormSet(request.POST or None, prefix='choices')

    if request.method == 'POST':
        q_type = request.POST.get('question_type', 'single')

        if q_type in ('true_false', 'essay'):
            # Đúng/Sai và Tự luận không dùng formset đáp án
            if form.is_valid():
                question = form.save(commit=False)
                question.created_by = request.user
                question.save()
                if q_type == 'true_false':
                    _save_true_false_choices(question, request.POST.get('tf_correct', 'true'))
                messages.success(request, 'Tạo câu hỏi thành công!')
                return redirect('question_bank')
            messages.error(request, 'Vui lòng kiểm tra lại các trường bị lỗi bên dưới.')
        else:
            if form.is_valid() and formset.is_valid():
                err = _validate_choices(formset, q_type)
                if err:
                    messages.error(request, err)
                else:
                    question = form.save(commit=False)
                    question.created_by = request.user
                    question.save()
                    order = 0
                    for f in formset:
                        cd = getattr(f, 'cleaned_data', None) or {}
                        if cd.get('DELETE') or not (cd.get('content') or '').strip():
                            continue
                        order += 1
                        Choice.objects.create(
                            question=question,
                            label=(cd.get('label') or '').strip() or chr(64 + order),
                            content=cd['content'].strip(),
                            is_correct=cd.get('is_correct', False),
                            order=order,
                        )
                    messages.success(request, 'Tạo câu hỏi thành công!')
                    return redirect('question_bank')
            else:
                messages.error(request, 'Vui lòng kiểm tra lại các trường bị lỗi bên dưới.')

    return render(request, 'questions/question_form.html', {
        'form': form, 'formset': formset, 'title': 'Tạo câu hỏi mới', 'tf_correct': 'true',
    })


@teacher_required
def question_edit_view(request, pk):
    question = get_object_or_404(Question, pk=pk)
    form = QuestionForm(request.POST or None, request.FILES or None, instance=question)
    formset = ChoiceFormSet(request.POST or None, instance=question, prefix='choices')

    # Đáp án đúng hiện tại của câu Đúng/Sai (cho radio khi edit)
    tf_correct = 'true'
    if question.question_type == 'true_false':
        correct = question.choices.filter(is_correct=True).first()
        if correct and correct.content.strip().lower() not in ('đúng', 'dung', 'true'):
            tf_correct = 'false'

    if request.method == 'POST':
        q_type = request.POST.get('question_type', question.question_type)

        if q_type in ('true_false', 'essay'):
            if form.is_valid():
                question = form.save()
                if q_type == 'true_false':
                    _save_true_false_choices(question, request.POST.get('tf_correct', 'true'))
                else:
                    question.choices.all().delete()
                messages.success(request, 'Cập nhật câu hỏi thành công!')
                return redirect('question_bank')
            messages.error(request, 'Vui lòng kiểm tra lại các trường bị lỗi bên dưới.')
        else:
            if form.is_valid() and formset.is_valid():
                err = _validate_choices(formset, q_type)
                if err:
                    messages.error(request, err)
                else:
                    form.save()
                    choices = formset.save(commit=False)
                    for choice in choices:
                        choice.question = question
                        choice.save()
                    for obj in formset.deleted_objects:
                        obj.delete()
                    # Đánh lại thứ tự
                    for i, c in enumerate(question.choices.order_by('order', 'id'), 1):
                        if c.order != i:
                            c.order = i
                            c.save(update_fields=['order'])
                    messages.success(request, 'Cập nhật câu hỏi thành công!')
                    return redirect('question_bank')
            else:
                messages.error(request, 'Vui lòng kiểm tra lại các trường bị lỗi bên dưới.')

    return render(request, 'questions/question_form.html', {
        'form': form, 'formset': formset, 'title': 'Chỉnh sửa câu hỏi', 'obj': question,
        'tf_correct': tf_correct,
    })


@teacher_required
def question_delete_view(request, pk):
    question = get_object_or_404(Question, pk=pk)
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Đã xóa câu hỏi.')
        return redirect('question_bank')
    return render(request, 'questions/question_confirm_delete.html', {'obj': question})


@teacher_required
def question_bulk_delete_view(request):
    """Xóa nhiều câu hỏi đã tick chọn trong ngân hàng câu hỏi."""
    if request.method != 'POST':
        return redirect('question_bank')
    ids = [i for i in request.POST.getlist('question_ids') if i.isdigit()]
    if not ids:
        messages.warning(request, 'Bạn chưa chọn câu hỏi nào để xóa.')
    else:
        target = Question.objects.filter(id__in=ids)
        count = target.count()
        target.delete()
        messages.success(request, f'Đã xóa {count} câu hỏi khỏi ngân hàng.')
    # Quay lại đúng trang/bộ lọc trước đó
    params = request.POST.get('return_query', '')
    url = '/questions/'
    if params:
        url += '?' + params
    return redirect(url)


@teacher_required
def import_questions_view(request):
    form = ImportQuestionForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        file = request.FILES['file']
        subject = form.cleaned_data['subject']
        filename = file.name.lower()
        try:
            if filename.endswith('.xlsx'):
                count = import_questions_from_excel(file, subject, request.user)
            elif filename.endswith('.docx'):
                count = import_questions_from_docx(file, subject, request.user)
            else:
                messages.error(request, 'Định dạng file không hợp lệ. Chỉ hỗ trợ .xlsx và .docx')
                return redirect('import_questions')
            messages.success(request, f'Import thành công {count} câu hỏi!')
        except Exception as e:
            messages.error(request, f'Lỗi khi import: {str(e)}')
        return redirect('question_bank')
    return render(request, 'questions/import.html', {'form': form})


@teacher_required
def export_questions_view(request):
    subject_id = request.GET.get('subject', '')
    qs = Question.objects.select_related('subject', 'topic').prefetch_related('choices')
    if subject_id:
        qs = qs.filter(subject_id=subject_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Questions'
    headers = ['STT', 'Môn học', 'Chủ đề', 'Loại', 'Độ khó', 'Điểm', 'Nội dung câu hỏi', 'A', 'B', 'C', 'D', 'Đáp án đúng', 'Giải thích']
    ws.append(headers)

    for i, q in enumerate(qs, 1):
        choices = list(q.choices.order_by('order'))
        correct_labels = ','.join(c.label for c in choices if c.is_correct)
        row = [i, q.subject.name, q.topic.name if q.topic else '',
               q.get_question_type_display(), q.get_difficulty_display(), float(q.points), q.content]
        for j in range(4):
            row.append(choices[j].content if j < len(choices) else '')
        row.extend([correct_labels, q.explanation])
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="question_bank.xlsx"'
    wb.save(response)
    return response


@teacher_required
def subject_list_view(request):
    subjects = Subject.objects.annotate(question_count=Count('questions')).order_by('name')
    return render(request, 'questions/subject_list.html', {'subjects': subjects})


@teacher_required
def subject_create_view(request):
    form = SubjectForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        subject = form.save(commit=False)
        subject.created_by = request.user
        subject.save()
        messages.success(request, 'Tạo môn học thành công!')
        return redirect('subject_list')
    return render(request, 'questions/subject_form.html', {'form': form})


def get_topics_ajax(request):
    subject_id = request.GET.get('subject_id')
    topics = Topic.objects.filter(subject_id=subject_id).order_by('name').values('id', 'name')
    return JsonResponse({'topics': list(topics)})


@teacher_required
def topic_create_ajax(request):
    """Tạo chủ đề mới cho một môn học (gọi từ modal ở form câu hỏi)."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Chỉ hỗ trợ POST'}, status=405)
    subject_id = request.POST.get('subject_id')
    name = (request.POST.get('name') or '').strip()
    description = (request.POST.get('description') or '').strip()
    if not subject_id or not name:
        return JsonResponse({'status': 'error', 'message': 'Vui lòng chọn môn học và nhập tên chủ đề.'}, status=400)
    subject = Subject.objects.filter(pk=subject_id).first()
    if not subject:
        return JsonResponse({'status': 'error', 'message': 'Môn học không tồn tại.'}, status=404)
    topic, created = Topic.objects.get_or_create(
        subject=subject, name=name, defaults={'description': description}
    )
    return JsonResponse({
        'status': 'created' if created else 'exists',
        'topic': {'id': topic.id, 'name': topic.name},
        'message': 'Tạo chủ đề thành công!' if created else 'Chủ đề này đã tồn tại trong môn học.',
    })


@teacher_required
def download_import_template_view(request):
    """Tải file Excel mẫu đúng định dạng import (kèm 2 dòng ví dụ)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cau hoi'
    headers = [
        'Nội dung câu hỏi', 'Loại', 'Độ khó',
        'Đáp án A', 'Đáp án B', 'Đáp án C', 'Đáp án D', 'Đáp án E',
        'Đáp án đúng', 'Giải thích', 'Môn học', 'Chủ đề', 'Điểm',
    ]
    ws.append(headers)
    ws.append([
        'Python là ngôn ngữ thông dịch?', 'true_false', 'easy',
        'Đúng', 'Sai', '', '', '',
        'A', 'Python là ngôn ngữ thông dịch (interpreted).',
        'Lập trình Python', 'Cơ bản', 1,
    ])
    ws.append([
        'Những kiểu dữ liệu nào là mutable?', 'multiple', 'medium',
        'list', 'dict', 'tuple', 'str', '',
        'A,B', 'list và dict thay đổi được; tuple, str thì không.',
        'Lập trình Python', 'Kiểu dữ liệu', 2,
    ])
    # Độ rộng cột dễ nhìn
    widths = [45, 12, 10, 18, 18, 18, 18, 12, 12, 40, 18, 15, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mau_import_cau_hoi.xlsx"'
    wb.save(response)
    return response
